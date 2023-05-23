import openpyxl
import re

#Open Excel workbook that's in source folder
#or replace with filepath to file, for example:
# wb = openpyxl.load_workbook(r'C:\Users\ericl\Desktop\intern\files\uoce sections.xlsx')
wb = openpyxl.load_workbook('uoce sections.xlsx')

#Open worksheet
ws = wb['Sheet1']
# ws = wb.worksheets[0]
sectionList = []
otherList = []

def version_key(x):
    return [int(i) for i in x.split('.')]


#Reiterate every row
for x in range(2, ws.max_row+1):
    #Get the cell value
    cell = ws.cell(x, column=1).value

    #Splitting the cell into two sections and get the number only.
    #(Section 301.1 = 'Section', '301.1')
    section = cell.split(" ",1)
    sectionNumber = section[1]
    # sectionNumber = re.sub('[a-zA-Z]','',sectionNumber)

    #If it is a number, add it to sectionList
    if sectionNumber is not None:
        #If section number contains non-numeric characters, add to other list.
        if re.search('[@_!#$%^&*()<>?/|}{~:]', sectionNumber) or re.search('[a-zA-Z]', sectionNumber):
            otherList.append(sectionNumber)
        else:
            sectionList.append(sectionNumber)

# print(otherList)
# print(sorted(sectionList, key=version_key))
sortedList = sorted(sectionList, key=version_key)

#Create new sheet called 'Sorted'
wb.create_sheet('Sorted')
ws = wb['Sorted']

#Column for unsorted section numbers that contains non-numericals such as S101 or 101.1(2)
ws.cell(row=1,column=1).value = "Unsorted List"
startingrow = 2
for x in otherList:
    ws.cell(startingrow, column=1).value = "Section " + x
    startingrow += 1

#Column for sorted section numbers
ws.cell(row=1, column=2).value = "Sorted List"
startingrow = 2
for x in sortedList:
    ws.cell(startingrow, column=2).value = "Section " + x
    startingrow +=1

wb.save('uoce sections.xlsx')